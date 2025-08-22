# exports.py
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from .models import CargoPostulacion, VotoMesaCargo
from django.db.models import Sum

@login_required
def export_summary_excel(request):
    cargo_param = (request.GET.get('cargo') or '').strip()
    cargo_id    = request.GET.get('cargo_id')

    cargo = None
    # 1) por id (si viene)
    if cargo_id and cargo_id.isdigit():
        cargo = CargoPostulacion.objects.filter(id=int(cargo_id)).first()

    # 2) por nombre con aliases y case-insensitive
    if not cargo and cargo_param:
        aliases = {
            'DIPUTADOS': 'Diputados Provinciales',
            'DIPUTADOS PROVINCIALES': 'Diputados Provinciales',
            'CONCEJALES': 'Concejales',
            'CONCEJAL': 'Concejales',
        }
        wanted = aliases.get(cargo_param.upper(), cargo_param)
        cargo = CargoPostulacion.objects.filter(nombre_postulacion__iexact=wanted).first()

    if not cargo:
        return HttpResponse("Cargo inválido", status=400)

    qs = (VotoMesaCargo.objects
          .select_related('partido_postulacion__partido')
          .filter(partido_postulacion__cargo_postulacion=cargo)
          .values('partido_postulacion__partido__nombre_partido')
          .annotate(votos=Sum('votos'))
          .order_by('-votos'))

    total_validos = sum((row['votos'] or 0) for row in qs)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Resumen {cargo.nombre_postulacion}"

    ws.append([f"Resumen {cargo.nombre_postulacion}"])
    ws.append([])
    ws.append(["Partido", "Votos", "%"])

    for row in qs:
        votos = row['votos'] or 0
        pct = (votos * 100 / total_validos) if total_validos else 0
        ws.append([row['partido_postulacion__partido__nombre_partido'], votos, round(pct, 2)])

    # ancho de columnas
    for idx, width in enumerate([40, 12, 10], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = cargo.nombre_postulacion.replace(' ', '_')
    resp['Content-Disposition'] = f'attachment; filename=Resumen_{safe_name}.xlsx'
    wb.save(resp)
    return resp
